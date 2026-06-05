from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import pandas as pd
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
import requests
import urllib.parse
import re
import mimetypes
import threading

app = Flask(__name__)
CORS(app)

BASE_DESTINATION = 'extracted_resumes'
UPLOAD_FOLDER = 'uploads'
extraction_status = {
    'status': 'idle',
    'message': '',
    'progress': 0,
    'total': 0,
    'extracted': 0,
    'errors': [],
    'last_file': None,
}

def find_resume_column(df):
    """Auto-detect the resume column"""
    common_patterns = ['resume', 'link', 'path', 'url', 'file']
    
    for col in df.columns:
        col_lower = col.lower()
        if any(pattern in col_lower for pattern in common_patterns):
            return col
    
    return df.columns[0]

def extract_resumes_from_file(excel_file):
    """Extract resumes from a single Excel file"""
    global extraction_status
    file_name_without_ext = Path(excel_file).stem
    destination_folder = os.path.join(BASE_DESTINATION, file_name_without_ext)

    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    try:
        wb = load_workbook(excel_file, read_only=False, data_only=False)
        ws = wb.active
    except Exception as e:
        extraction_status['message'] = f"Error reading {excel_file}: {e}"
        return 0, [(None, None, f"Error reading file: {e}")]

    keywords = ['resume', 'link', 'path', 'url', 'file']
    header_row_idx = None
    max_scan = min(12, ws.max_row)
    for r in range(1, max_scan + 1):
        row_vals = [str(c.value).lower() if c.value is not None else '' for c in ws[r]]
        if any(any(k in v for k in keywords) for v in row_vals):
            header_row_idx = r
            break

    if header_row_idx is None:
        header_row_idx = 1

    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[header_row_idx]]
    col_index = 0
    for i, h in enumerate(headers):
        hl = h.lower()
        if any(p in hl for p in keywords):
            col_index = i
            break

    count = 0
    errors = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        extraction_status['progress'] = r
        cell = ws.cell(row=r, column=col_index + 1)

        target = None
        if getattr(cell, 'hyperlink', None) and getattr(cell.hyperlink, 'target', None):
            target = cell.hyperlink.target
        elif cell.value and isinstance(cell.value, str):
            target = cell.value.strip()

        if not target:
            errors.append((r, None, 'no link/path found'))
            continue

        if target.startswith('file://'):
            parsed = urllib.parse.urlparse(target)
            local_path = urllib.parse.unquote(parsed.path)
            if os.name == 'nt' and local_path.startswith('/') and ':' in local_path:
                local_path = local_path.lstrip('/')
            source = local_path
            is_url = False
        elif target.startswith('http://') or target.startswith('https://'):
            source = target
            is_url = True
        else:
            source = os.path.expanduser(target)
            if not os.path.isabs(source):
                source = os.path.join(os.path.dirname(excel_file), source)
            is_url = False

        try:
            if is_url:
                resp = requests.get(source, stream=True, timeout=30)
                if resp.status_code == 200:
                    cd = resp.headers.get('content-disposition') or resp.headers.get('Content-Disposition')
                    file_name = None
                    if cd:
                        m = re.search(r'filename\*?=([^;]+)', cd)
                        if m:
                            file_name = m.group(1).strip().strip('"')
                    if not file_name:
                        parsed = urllib.parse.urlparse(source)
                        file_name = os.path.basename(parsed.path) or f"resume_{r}"

                    file_name = urllib.parse.unquote(file_name)
                    file_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', file_name)
                    file_name = file_name.strip()
                    if len(file_name) > 200:
                        ext = os.path.splitext(file_name)[1]
                        file_name = file_name[:200 - len(ext)] + ext

                    if '.' not in file_name:
                        ctype = resp.headers.get('content-type', '')
                        ext = mimetypes.guess_extension(ctype.split(';')[0].strip()) if ctype else None
                        if ext:
                            file_name += ext

                    file_name = file_name.lstrip('/\\')
                    dest_path = os.path.join(destination_folder, file_name)
                    with open(dest_path, 'wb') as f:
                        for chunk in resp.iter_content(chunk_size=8192):
                            if chunk:
                                f.write(chunk)
                    count += 1
                    extraction_status['extracted'] = count
                else:
                    reason = f"HTTP {resp.status_code}"
                    errors.append((r, source, reason))
            else:
                if os.path.exists(source):
                    file_name = os.path.basename(source)
                    dest_path = os.path.join(destination_folder, file_name)
                    shutil.copy2(source, dest_path)
                    count += 1
                    extraction_status['extracted'] = count
                else:
                    errors.append((r, source, 'local file not found'))
        except Exception as e:
            errors.append((r, source, str(e)))

    return count, errors

def process_files(files):
    """Process multiple Excel files"""
    global extraction_status
    
    extraction_status = {
        'status': 'processing',
        'message': f'Processing {len(files)} files...',
        'progress': 0,
        'total': 0,
        'extracted': 0,
        'errors': [],
        'last_file': None,
    }
    
    total_count = 0
    all_errors = []
    for excel_file in files:
        extraction_status['message'] = f'Processing: {excel_file}'
        extraction_status['last_file'] = excel_file
        count, errors = extract_resumes_from_file(excel_file)
        total_count += count
        all_errors.extend([{'row': row, 'source': source, 'reason': reason, 'file': excel_file} for row, source, reason in errors])
    
    extraction_status['status'] = 'completed'
    extraction_status['message'] = f'All tasks complete. {total_count} resumes extracted.'
    extraction_status['extracted'] = total_count
    extraction_status['errors'] = all_errors

@app.route('/api/upload', methods=['POST'])
def upload_files():
    """Handle file uploads"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    file_paths = []
    
    for file in files:
        if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            file_path = os.path.join(UPLOAD_FOLDER, file.filename)
            file.save(file_path)
            file_paths.append(file_path)
    
    if not file_paths:
        return jsonify({'error': 'No valid Excel files provided'}), 400
    
    # Start processing in background thread
    thread = threading.Thread(target=process_files, args=(file_paths,))
    thread.daemon = True
    thread.start()
    
    return jsonify({'message': 'Processing started', 'files': file_paths}), 202

@app.route('/api/status', methods=['GET'])
def get_status():
    """Get extraction status"""
    return jsonify(extraction_status), 200

@app.route('/api/results', methods=['GET'])
def get_results():
    """Get list of extracted resumes"""
    results = {}
    
    if os.path.exists(BASE_DESTINATION):
        for folder in os.listdir(BASE_DESTINATION):
            folder_path = os.path.join(BASE_DESTINATION, folder)
            if os.path.isdir(folder_path):
                files = os.listdir(folder_path)
                results[folder] = files
    
    return jsonify(results), 200


@app.route('/')
def serve_ui_root():
    """Serve the frontend index.html at the root path"""
    try:
        return send_from_directory('.', 'index.html')
    except Exception:
        return jsonify({'error': 'UI not found. Place index.html in the project root.'}), 404


if __name__ == '__main__':
    app.run(debug=True, port=5000)
